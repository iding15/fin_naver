import xlrd
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

class CheckBox:
    def __init__(self, name):
        self.name = name
        self.path = './DATA/'+self.name+'.xlsx'
        self.df = pd.read_excel(self.path, engine='openpyxl')

    def delNan(self):
        df = self.df.dropna(how='all')
        df = df.dropna(axis = 1, how='all')
        return df

def save_sheet(name, df, sh_name):
    path = './DATA/'+name+'.xlsx'
    writer = pd.ExcelWriter(path, engine='openpyxl') # pylint: disable=abstract-class-instantiated

    book = load_workbook(path)
    writer.book = book

    df.to_excel(writer, sheet_name=sh_name)
    writer.save()
    writer.close()
def remove_sheet(name, sh_name):
    path = './DATA/'+name+'.xlsx'
    wb = load_workbook(path)
    del wb[sh_name]
    wb.save(path)
